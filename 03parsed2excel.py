# This code is designed to take easyOCR parsed file with a list of 5 element lists (Lap number, total distance in meters, minutes, seconds, strokes/split(lap) and writes them into an excel workbook that analyzes the output of easyOCRtest and output a list with a list containing each row of data run the code by typing python easyOCR2excel.py <inputfilename> 
#  python easyOCR2excel.py <####outparsed_out> <worksheet title> 
import sys  # Used to read commandline arguments
import ast 
import openpyxl 
import re

# Load the existing Excel file
file_path = "workouts.xlsx"  # Replace with the path to your Excel file
wb = openpyxl.load_workbook(file_path)

# How to use commandline arguments: python 03parsed2excel.py <dateparsed>
in_file=sys.argv[1]

# Removes characters in the input file title that are not numbers to name the worksheet by date
# print("Removing all the characters except digits")
ws_name=re.sub("\D", "", in_file)
ws = wb.create_sheet(ws_name, 1) #Insert worksheet in 2nd position
wb.active = wb[ws_name]
# print(wb.sheetnames)

# Initialize an empty list to store the read lines
in_list = []
# Open the file for reading
with open(in_file, 'r') as file:
    for line in file:
        # Remove trailing newline characters and add the line to the list
        in_list.append(line.strip())      

# Convert the string representations of lists to actual lists
list_of_lists = [ast.literal_eval(item) for item in in_list]
splits = len(list_of_lists)
fdr = 4 #first data row
row = fdr-1

# Determine if laps are 25 or 50 meters
laplength = int(list_of_lists[0][1].strip('m')) 

# Write data to worksheet
for row_num in list_of_lists:
    if laplength == 25:
        row += 1
        line_num = int(row_num[0])
        ws.cell(row=row, column=1).value = line_num
        rundis = int(row_num[1].strip('m'))
        ws.cell(row=row, column=2).value = rundis
        laptime=int(row_num[2])*60+int(row_num[3])
        ws.cell(row=row, column=3).value = laptime
        stroke=int(row_num[4])
        ws.cell(row=row, column=4).value = stroke
        # print(str(line_num) +" "+ str(rundis) + " " + str(laptime)+" "+str(stroke))
        ldr = splits + fdr - 1 #last data row
    elif laplength == 50:
        row += 1
        line_num=row-fdr+1
        ws.cell(row=row, column=1).value = line_num
        rundis = line_num*25
        ws.cell(row=row, column=2).value = rundis
        laptime=int(((int(row_num[2])*60)+int(row_num[3]))/2)
        ws.cell(row=row, column=3).value = laptime
        stroke = int(int(row_num[4])/2)
        ws.cell(row=row, column=4).value = stroke
        # print(str(line_num) +" "+ str(rundis) + " " + str(laptime)+" "+str(stroke)) 
        row += 1
        line_num=row-fdr+1
        ws.cell(row=row, column=1).value = line_num
        rundis = line_num*25
        ws.cell(row=row, column=2).value = rundis
        laptime=int(((int(row_num[2])*60)+int(row_num[3]))/2)
        extratime=int(((int(row_num[2])*60)+int(row_num[3]))%2)
        laptime = laptime + extratime
        ws.cell(row=row, column=3).value = laptime
        stroke = int(int(row_num[4])/2)+int(int(row_num[4])%2)
        ws.cell(row=row, column=4).value = stroke
        ldr = splits*2 + fdr - 1 #last data row 
        # print(str(line_num) +" "+ str(rundis) + " " + str(laptime)+" "+str(stroke))   

# Write column headers
row = 1
ws.cell(row=row, column=1).value = '#splits'
ws.cell(row=row, column=2).value = 'Tot mtrs'
ws.cell(row=row, column=3).value = 25
ws.cell(row=row, column=4).value = 'strokes'

# First use the total distance to find the last length to use
total_distance = ws.cell(row=ldr, column=2).value
pos_distance = [50,100,250,450,500,800,1000,1600,2000,3225,4000,4025,4825,5000,5625,6000,6450,7000,7250,8000,8050,8850,9000,9650,10000,10450,11000,11250,12000,12050]
distance=[]
for i in pos_distance:
    if (int(i) < total_distance):
        distance.append(i)
        
for i,line in enumerate(distance,0):
    ws.cell(row=row, column=i+5).value = distance[i]
 
# Adding the formulas to calculate statistics
# Write Formula row
# Function to convert numeric column number n into excel column letter for use in formulas
def n2ec(n):
    result = ''
    while n > 0:
        n -= 1
        quotient, remainder = divmod(n, 26)
        result = chr(65 + remainder) + result  # 65 is the ASCII code for 'A'
        n = quotient
    return result

# Write the formulas for row 2
row = 2
col = 1
ec = n2ec(col)
ws.cell(row=row, column=col).value = '='+ec+str(ldr)
col = 2
ec = n2ec(col)
ws.cell(row=row, column=col).value = '='+ec+str(ldr)
col = 3
ec = n2ec(col)
min_formula= '=min('+ec+str(fdr)+':'+ec+str(ldr)+')'
ws.cell(row=row, column=col).value = min_formula
col = 4
ec = n2ec(col)
min_formula= '=min('+ec+str(fdr)+':'+ec+str(ldr)+')'
ws.cell(row=row, column=col).value = min_formula

# Write the formulas from row 3
row = 3 
col = 3
ec = n2ec(col)
avg_formula= '=int(trimmean('+ec+str(fdr)+':'+ec+str(ldr)+',0.1))'
ws.cell(row=row, column=col).value = avg_formula
col = 4
ec = n2ec(col)
avg_formula= '=int(trimmean('+ec+str(fdr)+':'+ec+str(ldr)+',0.1))'
ws.cell(row=row, column=col).value = avg_formula

# Find the best and average for each distance from the distance table
# Write the header
rcol = 3       #The time column data is being read from
ec = n2ec(rcol)

wcol = 5
for i in distance:
    wec = n2ec(wcol)
    sum_rows = int(i/25)
    row = fdr + sum_rows -1
    min_formula= '=min('+wec+str(row)+':'+wec+str(ldr)+')'
    ws.cell(row=2, column=wcol).value = min_formula
    avg_formula= '=int(trimmean('+wec+str(row)+':'+wec+str(ldr)+',0.1))'
    ws.cell(row=3, column=wcol).value = avg_formula
    
    while (row < ldr+1 ):
        sum_row = '=sum('+ec+str(row-sum_rows+1)+':'+ec+str(row)+')'
        ws.cell(row=row, column=wcol).value = sum_row
        row += 1
    wcol += 1
    
# Test Print
print(list(ws.values)) 
 
# save your new workbook!
wb.save("./workouts.xlsx")