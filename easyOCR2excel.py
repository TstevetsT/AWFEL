# This code is designed to take easyOCR parsed file with a list of 5 element lists (Lap number, total distance in meters, minutes, seconds, strokes/split(lap) and writes them into an excel workbook that analyzes the output of easyOCRtest and output a list with a list containing each row of data run the code by typing python easyOCR2excel.py <inputfilename> 
#  python easyOCR2excel.py 1960outparsed_out 
import sys  # Used to read commandline arguments
import ast 
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# Reading from file
in_file=sys.argv[1]

# Initialize an empty list to store the read lines
in_list = []

# Open the file for reading
with open(in_file, 'r') as file:
    for line in file:
        # Remove trailing newline characters and add the line to the list
        print(line)
        in_list.append(line.strip())      

# Convert the string representations of lists to actual lists
list_of_lists = [ast.literal_eval(item) for item in in_list]

# Write to a worksheet
row = 0
for row_num in list_of_lists:
    row += 1
    ws.cell(row=row, column=1).value = int(row_num[0])
    ws.cell(row=row, column=2).value = int(row_num[1].strip('m'))
    ws.cell(row=row, column=3).value = int(row_num[2])*60+int(row_num[3])
    ws.cell(row=row, column=4).value = int(row_num[4])
    
print(list(ws.values))
# save your new workbook!
wb.save("./workouts.xlsx")